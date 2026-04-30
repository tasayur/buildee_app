# excel_export.py -- BuildeeMgr Excel export engine (openpyxl)
import io
from datetime import date, datetime
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_DATE_YYYYMMDD2

import database as db

# ------------------------------------------------------------------
# Color palette
# ------------------------------------------------------------------
NAVY   = "1A1A2E"
ORANGE = "F97316"
GREEN  = "166534"
RED    = "DC2626"
GRAY   = "F1F5F9"
WHITE  = "FFFFFF"
LIGHT_ORANGE = "FFF7ED"
LIGHT_GREEN  = "DCFCE7"
LIGHT_RED    = "FEF2F2"
LIGHT_BLUE   = "DBEAFE"

# ------------------------------------------------------------------
# Style helpers
# ------------------------------------------------------------------
def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color=WHITE, size=10, name="Meiryo UI"):
    return Font(bold=bold, color=color, size=size, name=name)

def _border():
    s = Side(style="thin", color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _header_row(ws, row, values, widths=None):
    """Write a dark navy header row."""
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill    = _fill(NAVY)
        cell.font    = _font(bold=True)
        cell.border  = _border()
        cell.alignment = _align("center")
    if widths:
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w

def _data_cell(ws, row, col, value, bg=None, font_color="1A1A2E",
               bold=False, align="left", number_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill   = _fill(bg) if bg else _fill(WHITE)
    cell.font   = _font(bold=bold, color=font_color, size=10)
    cell.border = _border()
    cell.alignment = _align(align, wrap=(align == "left"))
    if number_format:
        cell.number_format = number_format
    return cell

def _title_block(ws, title, subtitle=""):
    """Rows 1-3: colored title block."""
    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 6
    t = ws.cell(row=1, column=1, value=title)
    t.fill = _fill(NAVY); t.font = Font(bold=True, color=WHITE, size=14, name="Meiryo UI")
    t.alignment = _align("left", "center")
    if subtitle:
        s = ws.cell(row=2, column=1, value=subtitle)
        s.fill = _fill(ORANGE); s.font = Font(bold=False, color=WHITE, size=9, name="Meiryo UI")
        s.alignment = _align("left", "center")

def _merge_title(ws, cols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=cols)

def _watermark_row(ws, row, cols):
    """Light gray info row at bottom."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1,
                value=f"BuildeeMgr  /  出力日時: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    c.font = Font(color="94A3B8", size=8, name="Meiryo UI", italic=True)
    c.alignment = _align("right")


# ==================================================================
# Sheet builders
# ==================================================================

def _sheet_schedules(wb, filter_date=None):
    ws = wb.create_sheet("作業予定・実績")
    rows = db.get_schedules(filter_date)

    sub = f"{filter_date} の作業予定" if filter_date else "全期間の作業予定"
    _title_block(ws, "作業予定・実績", sub)
    _merge_title(ws, 9)

    headers = ["日付", "協力会社", "作業内容", "作業場所", "人数", "開始", "終了", "備考", "ステータス"]
    widths  = [12, 20, 30, 18, 7, 8, 8, 24, 10]
    _header_row(ws, 4, headers, widths)

    STATUS_COLOR = {"完了": (LIGHT_GREEN, GREEN), "予定": (LIGHT_ORANGE, "C2410C")}
    for i, r in enumerate(rows, 5):
        bg_s, fg_s = STATUS_COLOR.get(r.get("status", ""), (GRAY, "475569"))
        stripe = GRAY if i % 2 == 0 else WHITE
        _data_cell(ws, i, 1, r.get("date", ""),         stripe)
        _data_cell(ws, i, 2, r.get("company", ""),      stripe)
        _data_cell(ws, i, 3, r.get("work_content", ""), stripe)
        _data_cell(ws, i, 4, r.get("location", ""),     stripe)
        _data_cell(ws, i, 5, r.get("workers_count", ""), stripe, align="center")
        _data_cell(ws, i, 6, r.get("time_start", ""),   stripe, align="center")
        _data_cell(ws, i, 7, r.get("time_end", ""),     stripe, align="center")
        _data_cell(ws, i, 8, r.get("note", ""),         stripe)
        _data_cell(ws, i, 9, r.get("status", ""),       bg_s, fg_s, bold=True, align="center")

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:I4"
    _watermark_row(ws, len(rows) + 6, 9)
    return ws


def _sheet_equipment(wb, filter_date=None):
    ws = wb.create_sheet("揚重機・重機予約")
    rows = db.get_equipment(filter_date)

    sub = f"{filter_date} の予約状況" if filter_date else "全期間の予約状況"
    _title_block(ws, "揚重機・重機予約", sub)
    _merge_title(ws, 7)

    headers = ["日付", "設備名", "協力会社", "開始", "終了", "使用時間(h)", "用途"]
    widths  = [12, 20, 20, 8, 8, 12, 28]
    _header_row(ws, 4, headers, widths)

    EQUIP_COLORS = {
        "タワークレーン":   "DBEAFE",
        "移動式クレーン":   "E0E7FF",
        "工事用エレベーター": "DCFCE7",
        "ゲート（北）":    "FEF9C3",
        "ゲート（南）":    "FFF7ED",
        "高所作業車":      "FCE7F3",
    }
    for i, r in enumerate(rows, 5):
        bg = EQUIP_COLORS.get(r.get("equipment", ""), GRAY)
        stripe = bg if i % 2 == 0 else WHITE

        # 使用時間を計算
        try:
            sh, sm = map(int, r["time_start"].split(":"))
            eh, em = map(int, r["time_end"].split(":"))
            hours = round(((eh * 60 + em) - (sh * 60 + sm)) / 60, 1)
        except Exception:
            hours = ""

        _data_cell(ws, i, 1, r.get("date", ""),       stripe)
        _data_cell(ws, i, 2, r.get("equipment", ""),  stripe, bold=True)
        _data_cell(ws, i, 3, r.get("company", ""),    stripe)
        _data_cell(ws, i, 4, r.get("time_start", ""), stripe, align="center")
        _data_cell(ws, i, 5, r.get("time_end", ""),   stripe, align="center")
        _data_cell(ws, i, 6, hours,                    stripe, align="center")
        _data_cell(ws, i, 7, r.get("purpose", ""),    stripe)

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = "A4:G4"
    _watermark_row(ws, len(rows) + 6, 7)
    return ws


def _sheet_ky(wb):
    ws = wb.create_sheet("電子KY記録")
    rows = db.get_ky_records()

    _title_block(ws, "電子KY（危険予知活動）記録", "全KY記録一覧")
    _merge_title(ws, 8)

    headers = ["日付", "協力会社", "作業内容", "危険ポイント", "対策・手順", "危険度", "ステータス", "承認日時"]
    widths  = [12, 20, 28, 36, 36, 8, 10, 18]
    _header_row(ws, 4, headers, widths)

    LEVEL_COLOR = {"高": (LIGHT_RED, RED), "中": (LIGHT_ORANGE, "C2410C"), "低": (LIGHT_GREEN, GREEN)}
    STATUS_COLOR = {"承認済": (LIGHT_GREEN, GREEN), "未承認": (LIGHT_ORANGE, "C2410C")}
    for i, r in enumerate(rows, 5):
        stripe = GRAY if i % 2 == 0 else WHITE
        lvl = r.get("level", "中")
        bg_l, fg_l = LEVEL_COLOR.get(lvl, (GRAY, "475569"))
        st = r.get("status", "未承認")
        bg_s, fg_s = STATUS_COLOR.get(st, (GRAY, "475569"))
        _data_cell(ws, i, 1, r.get("date", ""),         stripe)
        _data_cell(ws, i, 2, r.get("company", ""),      stripe)
        _data_cell(ws, i, 3, r.get("work_content", ""), stripe)
        _data_cell(ws, i, 4, r.get("danger_point", ""), stripe)
        _data_cell(ws, i, 5, r.get("measure", ""),      stripe)
        _data_cell(ws, i, 6, lvl, bg_l, fg_l, bold=True, align="center")
        _data_cell(ws, i, 7, st,  bg_s, fg_s, bold=True, align="center")
        _data_cell(ws, i, 8, r.get("approved_at", ""),  stripe, align="center")

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = "A4:H4"
    _watermark_row(ws, len(rows) + 6, 8)
    return ws


def _sheet_workers(wb):
    ws = wb.create_sheet("作業員名簿")
    rows = db.get_workers()
    today = date.today()

    _title_block(ws, "作業員名簿（グリーンファイル）", "登録作業員一覧")
    _merge_title(ws, 10)

    headers = ["氏名", "フリガナ", "協力会社", "職種", "生年月日", "血液型",
               "保険種別", "緊急連絡先", "資格名", "資格有効期限", "CCUS ID"]
    widths  = [14, 16, 20, 12, 13, 7, 20, 18, 20, 14, 18]
    # 11列
    _header_row(ws, 4, headers, widths)
    # merge title to 11 cols
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=11)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=11)

    for i, r in enumerate(rows, 5):
        stripe = GRAY if i % 2 == 0 else WHITE
        # 資格期限の色
        expiry = r.get("cert_expiry", "")
        bg_exp = stripe
        if expiry:
            try:
                exp = date.fromisoformat(expiry)
                days = (exp - today).days
                if days < 0:
                    bg_exp = LIGHT_RED
                elif days <= 30:
                    bg_exp = LIGHT_ORANGE
                else:
                    bg_exp = LIGHT_GREEN
            except Exception:
                pass

        _data_cell(ws, i,  1, r.get("name", ""),      stripe, bold=True)
        _data_cell(ws, i,  2, r.get("kana", ""),      stripe)
        _data_cell(ws, i,  3, r.get("company", ""),   stripe)
        _data_cell(ws, i,  4, r.get("job", ""),       stripe)
        _data_cell(ws, i,  5, r.get("birth", ""),     stripe, align="center")
        _data_cell(ws, i,  6, r.get("blood", ""),     stripe, align="center")
        _data_cell(ws, i,  7, r.get("insurance", ""), stripe)
        _data_cell(ws, i,  8, r.get("emergency", ""), stripe)
        _data_cell(ws, i,  9, r.get("cert_name", ""), stripe)
        _data_cell(ws, i, 10, expiry,                 bg_exp, align="center")
        _data_cell(ws, i, 11, r.get("ccus", ""),      stripe, align="center")

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = "A4:K4"
    _watermark_row(ws, len(rows) + 6, 11)
    return ws


def _sheet_attendance(wb, filter_date=None):
    ws = wb.create_sheet("入退場記録")
    rows = db.get_attendance(filter_date)

    sub = f"{filter_date} の入退場記録" if filter_date else "全期間の入退場記録"
    _title_block(ws, "入退場記録", sub)
    _merge_title(ws, 8)

    headers = ["日付", "氏名", "協力会社", "職種", "入場時刻", "退場時刻", "滞在時間(h)", "CCUS ID"]
    widths  = [12, 15, 20, 12, 10, 10, 12, 18]
    _header_row(ws, 4, headers, widths)

    # 集計
    total_present = 0
    for i, r in enumerate(rows, 5):
        stripe = GRAY if i % 2 == 0 else WHITE
        checkin  = r.get("checkin_time",  "")
        checkout = r.get("checkout_time", "")
        in_field = not checkout
        if in_field:
            total_present += 1
            stripe = LIGHT_GREEN  # 在場中は薄緑

        # 滞在時間計算
        duration = ""
        if checkin and checkout:
            try:
                sh, sm = map(int, checkin.split(":"))
                eh, em = map(int, checkout.split(":"))
                mins = (eh * 60 + em) - (sh * 60 + sm)
                duration = f"{mins // 60}h{mins % 60:02d}m"
            except Exception:
                pass
        elif in_field:
            duration = "在場中"

        name = r.get("worker_name") or r.get("worker_id", "")
        _data_cell(ws, i, 1, r.get("date", ""),    stripe)
        _data_cell(ws, i, 2, name,                  stripe, bold=True)
        _data_cell(ws, i, 3, r.get("company", ""), stripe)
        _data_cell(ws, i, 4, r.get("job", ""),     stripe)
        _data_cell(ws, i, 5, checkin,               stripe, align="center")
        _data_cell(ws, i, 6, checkout,              stripe, align="center")
        _data_cell(ws, i, 7, duration,
                   LIGHT_ORANGE if in_field else stripe, align="center")
        _data_cell(ws, i, 8, r.get("ccus", ""),   stripe, align="center")

    # サマリー行
    sr = len(rows) + 6
    ws.cell(row=sr, column=1, value="合計").font = Font(bold=True, name="Meiryo UI", size=10)
    ws.cell(row=sr, column=2, value=f"{len(rows)} 名入場").font = Font(bold=True, name="Meiryo UI", size=10)
    ws.cell(row=sr, column=3, value=f"在場中: {total_present} 名").font = Font(bold=True, name="Meiryo UI", size=10, color=GREEN)

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = "A4:H4"
    _watermark_row(ws, sr + 1, 8)
    return ws


def _sheet_safety_docs(wb):
    ws = wb.create_sheet("安全書類")
    rows = db.get_safety_docs()

    _title_block(ws, "安全書類一覧（グリーンファイル）", "提出済書類")
    _merge_title(ws, 5)

    headers = ["提出日", "書類種別", "協力会社", "備考", "ステータス"]
    widths  = [13, 30, 22, 30, 10]
    _header_row(ws, 4, headers, widths)

    for i, r in enumerate(rows, 5):
        stripe = GRAY if i % 2 == 0 else WHITE
        _data_cell(ws, i, 1, r.get("date", ""),      stripe, align="center")
        _data_cell(ws, i, 2, r.get("doc_type", ""),  stripe, bold=True)
        _data_cell(ws, i, 3, r.get("company", ""),   stripe)
        _data_cell(ws, i, 4, r.get("note", ""),      stripe)
        _data_cell(ws, i, 5, r.get("status", ""),    LIGHT_GREEN, GREEN,
                   bold=True, align="center")

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = "A4:E4"
    _watermark_row(ws, len(rows) + 6, 5)
    return ws


def _sheet_dashboard(wb):
    """1枚目: サマリーダッシュボードシート"""
    ws = wb.create_sheet("ダッシュボード", 0)  # 先頭に挿入
    today_str = date.today().isoformat()
    stats = db.get_dashboard_stats(today_str)
    expiring = db.get_expiring_certs(30)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 18

    # タイトル
    ws.merge_cells("A1:D1")
    t = ws.cell(row=1, column=1, value="BuildeeMgr 現場管理レポート")
    t.fill = _fill(NAVY); t.font = Font(bold=True, color=WHITE, size=16, name="Meiryo UI")
    t.alignment = _align("center", "center")
    ws.row_dimensions[1].height = 38

    ws.merge_cells("A2:D2")
    sub = ws.cell(row=2, column=1,
                  value=f"出力日: {today_str}  /  出力者: 現場管理者")
    sub.fill = _fill(ORANGE); sub.font = Font(color=WHITE, size=10, name="Meiryo UI")
    sub.alignment = _align("center", "center")
    ws.row_dimensions[2].height = 20

    # KPIブロック
    kpis = [
        ("本日の入場者数",    stats["today_workers"],    LIGHT_BLUE,  "1D4ED8"),
        ("本日の作業予定",    stats["today_schedules"],  LIGHT_GREEN, "166534"),
        ("未承認KY件数",     stats["pending_ky"],        LIGHT_ORANGE,"C2410C"),
        ("資格期限アラート",  stats["expiring_certs"],   LIGHT_RED,   "DC2626"),
    ]
    ws.row_dimensions[4].height = 14
    kpi_row = 5
    ws.row_dimensions[kpi_row].height = 52
    ws.row_dimensions[kpi_row + 1].height = 20
    for col, (label, val, bg, fg) in enumerate(kpis, 1):
        c_val = ws.cell(row=kpi_row,     column=col, value=val)
        c_lbl = ws.cell(row=kpi_row + 1, column=col, value=label)
        c_val.fill = _fill(bg); c_val.font = Font(bold=True, color=fg, size=28, name="Meiryo UI")
        c_val.alignment = _align("center", "center")
        c_lbl.fill = _fill(bg); c_lbl.font = Font(color="475569", size=10, name="Meiryo UI")
        c_lbl.alignment = _align("center", "center")
        # 枠線
        b = _border()
        c_val.border = b; c_lbl.border = b

    # 本日の作業予定テーブル
    ws.row_dimensions[8].height = 14
    ws.merge_cells("A9:D9")
    hdr = ws.cell(row=9, column=1, value="本日の作業予定")
    hdr.fill = _fill(NAVY); hdr.font = Font(bold=True, color=WHITE, size=11, name="Meiryo UI")
    hdr.alignment = _align("left", "center")
    ws.row_dimensions[9].height = 22

    _header_row(ws, 10, ["協力会社", "作業内容", "作業場所", "人数"])
    for i, s in enumerate(stats["recent_schedules"], 11):
        stripe = GRAY if i % 2 == 1 else WHITE
        _data_cell(ws, i, 1, s.get("company", ""),      stripe)
        _data_cell(ws, i, 2, s.get("work_content", ""), stripe)
        _data_cell(ws, i, 3, s.get("location", ""),     stripe)
        _data_cell(ws, i, 4, f"{s.get('workers_count',0)}名", stripe, align="center")
    if not stats["recent_schedules"]:
        ws.merge_cells("A11:D11")
        ws.cell(row=11, column=1, value="本日の作業予定はありません").font = \
            Font(color="94A3B8", italic=True, name="Meiryo UI")

    # 資格期限アラートテーブル
    start_row = 11 + max(len(stats["recent_schedules"]), 1) + 2
    ws.merge_cells(f"A{start_row}:D{start_row}")
    hdr2 = ws.cell(row=start_row, column=1, value="資格期限アラート（30日以内）")
    hdr2.fill = _fill(RED if expiring else NAVY)
    hdr2.font = Font(bold=True, color=WHITE, size=11, name="Meiryo UI")
    hdr2.alignment = _align("left", "center")
    ws.row_dimensions[start_row].height = 22

    _header_row(ws, start_row + 1, ["氏名", "資格名", "有効期限", "残日数"])
    for j, e in enumerate(expiring, start_row + 2):
        bg = LIGHT_RED if e["days"] <= 7 else LIGHT_ORANGE
        _data_cell(ws, j, 1, e["name"],   bg, bold=True)
        _data_cell(ws, j, 2, e["cert"],   bg)
        _data_cell(ws, j, 3, e["expiry"], bg, align="center")
        _data_cell(ws, j, 4, f'{e["days"]}日', bg,
                   RED if e["days"] <= 7 else "C2410C", bold=True, align="center")
    if not expiring:
        ws.merge_cells(f"A{start_row+2}:D{start_row+2}")
        ws.cell(row=start_row + 2, column=1,
                value="期限切れ間近の資格はありません").font = \
            Font(color="166534", name="Meiryo UI")

    _watermark_row(ws, start_row + max(len(expiring), 1) + 4, 4)
    return ws


# ==================================================================
# Public API — called from app.py
# ==================================================================

def build_excel(sheets: list, filter_date: str = None) -> bytes:
    """
    sheets: list of sheet names to include
            e.g. ["all"] or ["schedules","ky","workers"]
    Returns bytes of .xlsx file.
    """
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    all_sheets = sheets == ["all"] or "all" in sheets

    _sheet_dashboard(wb)   # always include dashboard summary

    if all_sheets or "schedules" in sheets:
        _sheet_schedules(wb, filter_date)
    if all_sheets or "equipment" in sheets:
        _sheet_equipment(wb, filter_date)
    if all_sheets or "ky" in sheets:
        _sheet_ky(wb)
    if all_sheets or "workers" in sheets:
        _sheet_workers(wb)
    if all_sheets or "attendance" in sheets:
        _sheet_attendance(wb, filter_date)
    if all_sheets or "safety_docs" in sheets:
        _sheet_safety_docs(wb)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
